from findElements import lerpdf, findDate, findPrice, findName
import csv
import xlsxwriter
import openpyxl

def print_comprovantes():
    for file_name in range(1, 100):
        try:
            text_list = lerpdf(str(file_name) + ".pdf")
            date = findDate(text_list, "2023")
            print(date)
            price = findPrice(text_list)
            print(price)
            name = findName(text_list)
            print(name)
            print("Documento " + str(file_name))
            print('_' * 10)
        except:
            if False:
                print()
    return comprovante

def read_comprovante():
    comprovante = [['DATA DO PAGAMENTO', 'VALOR', 'RAZÃO DA CONTA']]
    print(comprovante[0])
    for file_name in range(1, 100):
        try:
            text_list = lerpdf(str(file_name) + ".pdf")
            date = findDate(text_list, "2023")
            price = findPrice(text_list)
            name = findName(text_list)
            line = [date, price, name]
            print(line)
            comprovante.append(line)
        except:
            if False:
                print()

    return comprovante

def create_csv(file_name, sheet):
    with open(file_name + '.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(sheet)


def create_xlsx_comprovante(comprovante):
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook('comprovante.xlsx')
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0

    # Iterate over the data and write it out row by row.
    for date, price, name in comprovante:
        worksheet.write(row, col, date)
        worksheet.write(row, col + 1, price)
        worksheet.write(row, col + 2, name)
        row += 1

    workbook.close()

def create_xlsx_template(template):
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook('template2.xlsx')
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0
    print('@')
    # Iterate over the data and write it out row by row.
    for mes1, mes2, data, valor, razao, forma, comprovante, boleto, observacao, tag in template:
        print('@')
        worksheet.write(row, col, mes1)
        worksheet.write(row, col + 1, mes2)
        worksheet.write(row, col + 2, data)
        worksheet.write(row, col + 3, valor)
        worksheet.write(row, col + 4, razao)
        worksheet.write(row, col + 5, forma)
        worksheet.write(row, col + 6, comprovante)
        worksheet.write(row, col + 7, boleto)
        worksheet.write(row, col + 8, observacao)
        worksheet.write(row, col + 9, tag)
        row += 1

    workbook.close()

def read_template(comprovante, template_doc, max_col, max_line):
    wb_obj = openpyxl.load_workbook(template_doc + ".xlsx")
    sheet = wb_obj.active
    this_line = []
    template = [[]]
    for line in range(1, max_line + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=line, column=col)
            if cell.value == None:
                cell.value = ""
            this_line.append(cell.value)
        print(this_line)
        template.append(this_line)
        this_line = []
    return template

if __name__ == '__main__':
    #comprovante = print_comprovantes()
    #comprovante = read_comprovante()
    template = read_template('', 'template', 10, 43)
    #create_csv("comprovante", comprovante)
    #create_xlsx_comprovante(comprovante)
    create_xlsx_template(template)
    #comprovante = read_template("", "template")